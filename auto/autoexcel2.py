import openpyxl # openpyxl 라이브러리를 import합니다.
import os.path # os.path 라이브러리를 import합니다.

status = [10, 20, 30, 40] # 엑셀 파일에 저장할 데이터를 준비합니다.

excelFile = os.path.join('.', 'data2.xlsx') # 엑셀 파일의 경로를 지정합니다.

wb = openpyxl.load_workbook(excelFile) # 엑셀 파일을 엽니다.

ws = wb['IP 주소별 접속 통계'] # 'IP 주소별 접속 통계' 워크시트를 가져옵니다.

ws['A1'] = '이름' # '이름' 컬럼의 1행에 '이름'을 입력합니다.

ws.cell(row=2, column=1, value = '김태윤') # 2행, 1열에 '김태윤'을 입력합니다.

ws.append(status) # '상태' 컬럼의 마지막 행에 status 데이터를 추가합니다.

ws.cell(row=3, column=5, value='=SUM(A3:D3)') # '상태' 컬럼의 3행, 5열에 A3부터 D3까지의 합을 계산하여 입력합니다.

wb.save(excelFile) # 엑셀 파일을 저장합니다.

wb.close() # 워크북을 닫습니다.
