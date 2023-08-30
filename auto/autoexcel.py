import openpyxl
import os.path

# excelFile = os.path.join('.', 'data.xlsx')
# wb = openpyxl.load_workbook(excelFile)
# print(wb.sheetnames)
# 엑셀 파일 생성
wb = openpyxl.Workbook()

# 워크시트 생성
ws = wb.active
ws.title = 'IP 주소별 접속 통계'
wb.save('data2.xlsx')


wb.close()